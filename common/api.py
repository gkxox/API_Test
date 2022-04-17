# -*- coding: utf-8 -*- 
# @Time : 2022/4/14 21:49 
# @Author : wangbo

import json

import jsonpath
import requests
from openpyxl.styles import PatternFill

from utils.logutil import logger
from utils.util import DirPath

from utils.util import read_config

dir_path = DirPath()
config_path = dir_path.get_config_path()


class RunMethod:

    def __init__(self):
        self.login_url = read_config(config_path, 'Login', 'URL')
        self.login_xcb_url = read_config(config_path, 'Login_XCB', 'URL')

    # 定位登录方法，获取token
    def login(self, mobile='13800000002', password='123456'):

        url = self.login_url
        headers = {"Content-Type": "application/json;charset=UTF-8"}
        login_data = {"mobile": mobile, "password": password}
        res = requests.post(url=url, json=login_data, headers=headers).json()
        # 获取token
        token = 'Bearer ' + res['data']
        return token

    def login_xcb(self):

        url =self.login_xcb_url
        headers = {"Content-Type": "application/x-www-form-urlencoded"}
        login_xcb_data = "username=15895802308&passwordmd5=f50cbcad941c4feff765ad02fde12482"
        res = requests.post(url=url,headers=headers,data=login_xcb_data)
        cookie = res.cookies.get_dict()
        return cookie



        # 获取cookie
        # res = requests.post(url=self.login_url,json=login_data,headers=headers)
        # cookie = res.cookies.get_dict()
        # return cookie

    def do_get(self, url, headers,data):
        cookies = self.login_xcb()
        # 如果headers是str格式，转换为dict
        if isinstance(headers, str):
            headers = json.loads(headers)
        # 请求头添加token
        # headers['Authorization'] = self.login()
        # 如果data是str格式，转换为dict
        # if isinstance(data, str):
        #     data = json.loads(data)

        data = data.encode('utf-8')
        res = requests.get(url=url, params=data, headers=headers,cookies=cookies)
        return res

    def do_post(self, url, headers, data):
        cookies =self.login_xcb()
        # 如果headers是str格式，转换为dict
        if isinstance(headers, str):
            headers = json.loads(headers)
        # 请求头添加token
        headers['Authorization'] = self.login()
        # 如果data是dict格式，转换为str

        # if 'null' in data:
        #     data = json.loads(data)
        #
        # if isinstance(data, dict):
        #     data = json.dumps(data)
        # else:
        #     data = data.encode('utf-8')

        data = data.encode('utf-8')

        res = requests.post(url=url, data=data, headers=headers,cookies=cookies)
        return res

    def do_put(self, url, headers, data):
        # 如果headers是str格式，转换为dict
        if isinstance(headers, str):
            headers = json.loads(headers)
        # 请求头添加token
        headers['Authorization'] = self.login()
        # 如果data是str格式，转换为dict
        if 'null' in data:
            data = json.loads(data)

        if isinstance(data, dict):
            data = json.dumps(data)
        else:
            data = data.encode('utf-8')
        res = requests.put(url=url, data=data, headers=headers)

        return res

    def do_delete(self, url, headers, data):
        # 如果headers是str格式，转换为dict
        if isinstance(headers, str):
            headers = json.loads(headers)
        # 请求头添加token
        headers['Authorization'] = self.login()
        # 如果data是str格式，转换为dict
        if isinstance(data, str):
            data = json.loads(data)
        data = data.encode('utf-8')
        res = requests.delete(url=url, data=data, headers=headers)
        return res


    # 根据method调用对应的请求方法
    def run_main(self, method, url, headers, data):

        if method.lower() == 'get':
            res = self.do_get(url, headers, data)
            return res

        elif method.lower() == 'post':
            res = self.do_post(url, headers, data)
            return res

        elif method.lower() == 'put':
            res = self.do_put(url, headers, data)
            return res

        elif method.lower() == 'delete':
            res = self.do_delete(url, headers, data)
            return res

        else:
            print("请求方式不是get或者post")

    # 根据key，逐层获取json格式的值
    def get_text(self, res, key, index=0):
        # 将str转换为dict
        text = json.loads(res)
        # 使用jsonpath解析数据
        value = jsonpath.jsonpath(text, '$..{0}'.format(key))

        # 如果value存在
        if value:
            # 如果value长度为1,返回第一个值
            if len(value) == 1:
                return value[0]
            # 如果value长度不为1,返回一个列表
            return value[index]
        # 如果value不存在，返回None
        else:
            return None

    # 根据var list，逐层获取json格式的值
    def parse_res(self, var, res, i=0):

        res = res.json()
        # 判断变量var是否存在
        if not var:
            # 不存在直接返回res内容
            return res

        else:
            # 存在则获取数组第1个内容
            res = res.get(var[0])
            # 从数组中删除第1个内容
            del var[0]
            # 如果res数据是list，取第i个值，默认取第一个值
            if isinstance(res, list):
                res = res[i]
            # 递归
            return self.parse_res(var, res, i)

    # 结果验证方法
    def assert_response(self, ws, id, data, res, expected, actual):
        # 变量初始化为False
        is_pass = False
        # 异常处理，捕获assert抛出的异常，不直接抛出
        try:
            # 根据结果进行断言验证
            assert expected == actual
            # 打印信息
            logger.info(f"用例断言成功,预期值：{expected},实际值：{actual}")

            # 测试结果写入pass，并标为绿色
            ws.cell(int(id) + 1, 10).value = 'Pass'
            fill = PatternFill('solid', fgColor='33cc00')
            ws.cell(int(id) + 1, 10).fill = fill

            # 设置变量为True
            is_pass = True
        # 捕获异常
        except:
            # 设置变量为False
            is_pass = False
            # 打印日志
            logger.info(f"用例断言失败,预期值：{expected},实际值：{actual}")

            # 测试结果写入fail，并标为红色
            ws.cell(int(id) + 1, 10).value = 'Fail'
            fill = PatternFill('solid', fgColor='FF0000')
            ws.cell(int(id) + 1, 10).fill = fill

        # 无论是否出现异常，都执行下面内容代码
        finally:
            # 把结果更新到表格
            ws.cell(int(id) + 1, 6).value = data
            ws.cell(int(id) + 1, 9).value = res.text
            # 根据变量结果是True/False，进行断言验证，成功则通过，失败则未通过
            assert is_pass
        # 返回该变量结果
        return is_pass


if __name__ == '__main__':
    run_main = RunMethod()

    # url = 'http://xcbapi.xuechebu.com/BaoMingApi/EmpInfo/GetEmpInfoList'
    # # headers = {"Content-Type": "application/json"}
    #
    # data = 'HdIds=&userwd=null&ossdk=24&regionid=2&pageIndex=1&pageSize=6&userjd=null&ipaddress=192.168.0.104&os=an&osversion=7.0&imei=0b1d73ba2ca8f9f6&appversion=10.0.2&version=10.0.2&CxIds=&sort=1'
    # data = json.dumps(data)
    # print(data)
    # res = run_main.do_get(url,data).text

    print(run_main.login_xcb())

    # print(RunMethod().get_text(data, 'id'))
