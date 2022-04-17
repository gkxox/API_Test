# -*- coding: utf-8 -*- 
# @Time : 2022/4/14 21:50 
# @Author : wangbo

from configparser import ConfigParser
from openpyxl.styles import PatternFill
from string import Template
import os
import yaml
import openpyxl
import json


class DirPath:

    def __init__(self):
        # 获取文件的绝对路径
        self.abs_path = os.path.abspath(__file__)
        # 获取文件所在目录的上一级目录,也就是根目录
        self.project_path = os.path.dirname(os.path.dirname(self.abs_path))
        # 通过os.sep的方法来获取config目录的全路径
        self._conf_path = self.project_path + os.sep + "config.ini"
        # 通过os.sep的方法来获取log日志目录的全路径
        self._log_path = self.project_path + os.sep + "log"
        # 通过os.sep的方法来获取report报告目录的全路径
        self._report_path = self.project_path + os.sep + "report"
        # 通过os.sep的方法来获取data数据目录的全路径
        self._data_path = self.project_path + os.sep + "data"

    # 返回日志目录
    def get_log_path(self):
        return self._log_path

    # 返回报告目录
    def get_report_path(self):
        return self._report_path

    # 返回config目录
    def get_config_path(self):
        return self._conf_path

    # 返回data目录
    def get_data_path(self):
        return self._data_path


# 读取config.ini方法
def read_config(file, section, option):
    conf = ConfigParser()
    conf.read(file)
    return conf.get(section, option)


# 解析yaml文件
def parse_yaml(file, section):
    # 打开yaml文件
    with open(file, 'r', encoding='utf8') as f:
        # 加载yaml数据
        yaml_data = yaml.safe_load(f)
        data_temp = yaml_data[section]

        # 将数据存放在data中
        data = []
        for i in data_temp:
            temp = []
            for x in i.values():
                temp.append(x)
            data.append(temp)
        return data


# 解析xls文件
def parse_xls(xls_name, sheet_name):
    # 打开表格
    wb = openpyxl.load_workbook(xls_name)
    ws = wb[sheet_name]
    # 获取表格行、列数
    rows = ws.max_row
    cols = ws.max_column

    # 将数据存放在data中
    data = []
    for row in range(2, rows + 1):
        data_temp = []
        for col in range(1, cols - 1):
            value = ws.cell(row, col).value
            data_temp.append(value)
        data.append(data_temp)

    # 关闭表格
    wb.close()
    return data


# 将测试结果写入表格中
def insert_result(ws, id, expected, actual):
    # 如果预期结果与实际结果一致
    if eval(expected) == eval(actual):
        # 测试结果写入pass，并标为绿色
        ws.cell(int(id) + 1, 10).value = 'Pass'
        fill = PatternFill('solid', fgColor='33cc00')
        ws.cell(int(id) + 1, 10).fill = fill

    # 如果预期结果与实际结果不一致
    else:
        # 测试结果写入fail，并标为红色
        ws.cell(int(id) + 1, 10).value = 'Fail'
        fill = PatternFill('solid', fgColor='FF0000')
        ws.cell(int(id) + 1, 10).fill = fill


# 处理数据之间的关联
def pasre_relation(data, replace_data):
    if isinstance(data, dict):
        data = json.dumps(data)

    s = Template(data)
    # 将data中${foo}关联数据，替换成所需 ，replace_data={"foo":"test"}
    data = s.safe_substitute(replace_data)
    return data


if __name__ == '__main__':
    # file = r'C:\Users\gkxox\Desktop\自动化测试\config.ini'
    # print(read_config(file, 'Login', 'URL'))
    #
    # print(DirPath().get_config_path())
    # print(parse_yaml(r'C:\Users\gkxox\Desktop\自动化测试\data\test_data.yaml', 'test01'))
    data = {"id":"${id}","name":"ww"}
    replace_data = {"id": "1063705989926227968"}

    print(pasre_relation(data, replace_data))
